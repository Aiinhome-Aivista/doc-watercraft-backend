from flask import request, jsonify, send_from_directory
import pandas as pd
from datetime import datetime
import os

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from database.db_connection import get_db_connection

EXPORT_FOLDER = "exports"
os.makedirs(EXPORT_FOLDER, exist_ok=True)


def export_full_report():
    try:
        conn = get_db_connection()

        # ================= FILTERS =================
        party_id = request.args.get("party_id")
        vessel_id = request.args.get("vessel_id")
        start = request.args.get("period_start")
        end = request.args.get("period_end")

        conditions = []
        params = []

        if party_id:
            conditions.append("v.party_id = %s")
            params.append(party_id)

        if vessel_id:
            conditions.append("v.id = %s")
            params.append(vessel_id)

        if start and end:
            conditions.append("DATE(v.created_at) BETWEEN %s AND %s")
            params.extend([start, end])

        where = " AND ".join(conditions)
        if where:
            where = "WHERE " + where

        # ================= QUERY =================
        query = f"""
        SELECT 
    v.vessel_name,
    pm.party_name,

    ge.gate_in_datetime,
    ge.gate_out_datetime,
    vm.vehicle_no,

    wr.gross_weight,
    wr.tare_weight,
    (wr.gross_weight - wr.tare_weight) AS net_weight,

    co.start_datetime,
    co.end_datetime,

    -- BILLING
    bm.voucher_number,
    bm.bill_date,
    bm.total_bill_value,

    bd.activity_name,
    bd.qty,
    bd.rate,
    bd.amount,
    bd.gst_amount

FROM vessels v
LEFT JOIN party_masters pm ON pm.id = v.party_id
LEFT JOIN gate_entries ge ON ge.party_id = v.party_id
LEFT JOIN vehicle_master vm ON vm.id = ge.vehicle_id
LEFT JOIN wbin_records wr ON wr.gate_entry_id = ge.id
LEFT JOIN cargo_operations co ON co.vessel_id = v.id
LEFT JOIN bill_details bd ON bd.vessel_id = v.id
LEFT JOIN bill_main bm ON bm.id = bd.bill_main_id
        {where}
        """

        df = pd.read_sql(query, conn, params=params)

        # ================= COLUMN RENAME =================
        df.columns = [
    "Vessel",
    "Party",
    "Gate In",
    "Gate Out",
    "Vehicle",
    "Gross",
    "Tare",
    "Net",
    "Start",
    "End",

    # BILLING
    "Voucher",
    "Bill Date",
    "Total Bill",
    "Activity",
    "Qty",
    "Rate",
    "Amount",
    "GST"
]

        # ================= FILE =================
        file_name = f"REPORT_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(EXPORT_FOLDER, file_name)

        with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="REPORT", index=False, startrow=4)

            ws = writer.book["REPORT"]

            # ================= HEADER =================
            ws.merge_cells("A1:K1")
            ws["A1"] = df["Vessel"].iloc[0] if not df.empty else "VESSEL REPORT"
            ws["A1"].font = Font(size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:K2")
            ws["A2"] = "Daily Loading & Unloading Data"
            ws["A2"].alignment = Alignment(horizontal="center")

            # ================= STYLES =================
            header_font = Font(bold=True)
            center_align = Alignment(horizontal="center")
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin")
            )

            # Header styling
            for col in range(1, len(df.columns) + 1):
                cell = ws.cell(row=5, column=col)
                cell.font = header_font
                cell.alignment = center_align
                cell.border = thin_border

            # ================= HIGHLIGHT (LIKE YOUR SHEET) =================
            net_col = df.columns.get_loc("Net") + 1

            for row in range(6, ws.max_row + 1):
                ws.cell(row=row, column=net_col).fill = yellow_fill

            # ================= AUTO WIDTH =================
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)

                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass

                ws.column_dimensions[col_letter].width = max_length + 2

            # ================= TOTAL =================
            total_row = ws.max_row + 2
            ws[f"H{total_row}"] = "Total Net:"
            ws[f"I{total_row}"] = df["Net"].sum()

            ws[f"H{total_row}"].font = Font(bold=True)
            ws[f"I{total_row}"].font = Font(bold=True)

        conn.close()

        return jsonify({
            "success": True,
            "download_url": f"/api/v1/export/download/{file_name}"
        })

    except Exception as e:
        return jsonify({
            "success": False,
            "error": str(e)
        })


def download_export(filename):
    return send_from_directory(EXPORT_FOLDER, filename, as_attachment=True)


def export_vehicle_movement_report():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        # ================= FILTERS =================
        vessel_id = request.args.get("vessel_id")
        party_id = request.args.get("party_id")
        start_date = request.args.get("start_date")
        end_date = request.args.get("end_date")

        if start_date is None and end_date is None:
            today = datetime.now().strftime("%Y-%m-%d")
            start_date = today
            end_date = today

        # ================= QUERY (Matching vessel_controller.py exactly) =================
        base_query = """
            FROM gate_entries ge
            LEFT JOIN party_masters pm ON pm.id = ge.party_id
            LEFT JOIN vehicle_master vm ON vm.id = ge.vehicle_id
            LEFT JOIN cargo_operations co ON co.id = (
                SELECT c2.id FROM cargo_operations c2 WHERE c2.gate_entry_id = ge.id ORDER BY c2.id DESC LIMIT 1
            )
            LEFT JOIN vessels v ON v.id = co.vessel_id
            LEFT JOIN wbin_records wbi ON wbi.gate_entry_id = ge.id
            LEFT JOIN wbout_records wbo ON wbo.gate_entry_id = ge.id
        """

        conditions = []
        params = []

        if vessel_id:
            conditions.append("co.vessel_id = %s")
            params.append(vessel_id)

        if party_id:
            conditions.append("ge.party_id = %s")
            params.append(party_id)

        if start_date and end_date:
            conditions.append("DATE(ge.gate_in_datetime) BETWEEN %s AND %s")
            params.extend([start_date, end_date])
        elif start_date:
            conditions.append("DATE(ge.gate_in_datetime) >= %s")
            params.append(start_date)
        elif end_date:
            conditions.append("DATE(ge.gate_in_datetime) <= %s")
            params.append(end_date)

        if conditions:
            base_query += " WHERE " + " AND ".join(conditions)

        data_query = f"""
            SELECT 
                pm.party_name,
                ge.gate_in_datetime,
                ge.challan_invoice_no,
                vm.vehicle_no,
                vm.transporter_name,
                ge.driver_name,
                ge.driver_mob_no,
                ge.outside_payment_slip,
                ge.outside_gross_weight,
                ge.outside_tare_weight,
                ge.outside_net_weight,
                wbi.wbin_datetime,
                wbi.gross_weight AS wbin_gross_weight,
                wbi.tare_weight AS wbin_tare_weight,
                wbo.wbout_datetime,
                wbo.gross_weight AS wbout_gross_weight,
                wbo.tare_weight AS wbout_tare_weight,
                ge.gate_out_datetime,
                co.start_datetime AS cargo_start_datetime,
                co.compressor_no,
                co.end_datetime AS cargo_end_datetime,
                v.id AS vessel_id,
                v.vessel_name,
                v.berthing_datetime,
                v.sailing_datetime,
                v.survey_quantity
            {base_query}
            ORDER BY ge.gate_in_datetime ASC
        """

        cursor.execute(data_query, tuple(params))
        cols = [c[0] for c in cursor.description]
        rows = cursor.fetchall()

        # Group data by vessel name
        vessel_groups = {}
        for row in rows:
            d = dict(zip(cols, row))
            v_name = d.get("vessel_name") or "Unassociated"
            if v_name not in vessel_groups:
                vessel_groups[v_name] = {
                    "vessel_name": v_name,
                    "berthing_datetime": d.get("berthing_datetime"),
                    "sailing_datetime": d.get("sailing_datetime"),
                    "survey_quantity": d.get("survey_quantity"),
                    "rows": []
                }
            
            # Perform vehicle movement calculations
            wbin_gross = d.get("wbin_gross_weight")
            wbin_tare = d.get("wbin_tare_weight")
            wbout_gross = d.get("wbout_gross_weight")
            wbout_tare = d.get("wbout_tare_weight")
            outside_net = d.get("outside_net_weight")
            
            own_gross = wbin_gross if wbin_gross is not None else wbout_gross
            own_tare = wbin_tare if wbin_tare is not None else wbout_tare
            
            net_val = None
            try:
                if wbin_gross is not None and wbout_tare is not None:
                    net_val = abs(float(wbin_gross) - float(wbout_tare))
                elif wbout_gross is not None and wbin_tare is not None:
                    net_val = abs(float(wbout_gross) - float(wbin_tare))
            except (ValueError, TypeError):
                net_val = None

            if net_val is not None:
                net_weight = round(net_val, 3)
            else:
                if wbin_gross is None and wbin_tare is None and wbout_gross is None and wbout_tare is None:
                    try:
                        net_weight = round(float(outside_net), 3) if outside_net is not None else None
                    except (ValueError, TypeError):
                        net_weight = outside_net
                else:
                    net_weight = None

            try:
                gross_weight = round(float(own_gross), 3) if own_gross is not None else None
            except (ValueError, TypeError):
                gross_weight = own_gross

            try:
                tare_weight = round(float(own_tare), 3) if own_tare is not None else None
            except (ValueError, TypeError):
                tare_weight = own_tare

            gate_in = d.get("gate_in_datetime")
            gate_out = d.get("gate_out_datetime")
            waiting_hours = None
            if isinstance(gate_in, datetime) and isinstance(gate_out, datetime):
                diff = gate_out - gate_in
                waiting_hours = round(diff.total_seconds() / 3600.0)

            def fmt_dt(val):
                if isinstance(val, datetime):
                    return val.strftime("%d-%m-%y(%H:%M)")
                return val or "—"

            vessel_groups[v_name]["rows"].append({
                "party_name": d.get("party_name") or "—",
                "gate_in_datetime": fmt_dt(gate_in),
                "challan_invoice_no": d.get("challan_invoice_no") or "—",
                "vehicle_no": d.get("vehicle_no") or "—",
                "transporter_name": d.get("transporter_name") or "—",
                "driver_name": d.get("driver_name") or "—",
                "driver_mob_no": d.get("driver_mob_no") or "—",
                "outside_payment_slip": d.get("outside_payment_slip") or "—",
                "outside_gross_weight": float(d.get("outside_gross_weight")) if d.get("outside_gross_weight") is not None else None,
                "outside_tare_weight": float(d.get("outside_tare_weight")) if d.get("outside_tare_weight") is not None else None,
                "outside_net_weight": float(d.get("outside_net_weight")) if d.get("outside_net_weight") is not None else None,
                "wbin_datetime": fmt_dt(d.get("wbin_datetime")),
                "gross_weight": float(gross_weight) if gross_weight is not None else None,
                "wbout_datetime": fmt_dt(d.get("wbout_datetime")),
                "tare_weight": float(tare_weight) if tare_weight is not None else None,
                "gate_out_datetime": fmt_dt(gate_out),
                "net_weight": float(net_weight) if net_weight is not None else None,
                "waiting_hours": waiting_hours,
                "cargo_start_datetime": fmt_dt(d.get("cargo_start_datetime")),
                "compressor_no": d.get("compressor_no") or "—",
                "cargo_end_datetime": fmt_dt(d.get("cargo_end_datetime"))
            })

        # ================= FILE =================
        file_name = f"VEHICLE_REPORT_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(EXPORT_FOLDER, file_name)

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "VEHICLE_REPORT"

        # Apply gridlines visibility
        ws.views.sheetView[0].showGridLines = True

        # Styles
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )

        vessel_header_font = Font(name="Calibri", size=11, bold=True)
        vessel_header_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        table_header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        table_header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        net_font = Font(name="Calibri", size=11, bold=True, color="B25E00")

        headers_cols = [
            "SL. No.",
            "Consignor - (Party Name)",
            "(GATE)\nEntry Date & Time as per",
            "CHALLAN &\nINV. NO",
            "VEHICLE NO.",
            "Transporter Name",
            "Driver name",
            "Driver number",
            "Out\nWeighment Slip no",
            "Out gross\nWt",
            "Out tare wt",
            "Out nett wt",
            "IN Weighment\nDATE & Time",
            "Gross\nWeight",
            "OUT Weighment\nDATE & Time",
            "Tare Weight",
            "(GATE)\nOut Date & Time as per",
            "Net\nMaterial Qty",
            "Waiting\nHour 24",
            "BULKER Unloading\nStart Date & Time\nStart to VESSEL",
            "COMPRESSOR\nOR NO.",
            "BULKER Unloading\nComplete Date & Time\nStart to VESSEL"
        ]

        numeric_cols = {
            10: "0.00", # Out gross Wt
            11: "0.00", # Out tare wt
            12: "0.00", # Out nett wt
            14: "0.00", # Gross Weight
            16: "0.00", # Tare Weight
            18: "0.00", # Net Material Qty
            19: "0"     # Waiting Hour 24
        }

        current_row = 1

        # In case there's no data
        if not vessel_groups:
            # Write a simple empty state
            ws.merge_cells("A1:V1")
            ws["A1"] = "NO VEHICLE MOVEMENT RECORDS FOUND"
            ws["A1"].font = Font(name="Calibri", size=14, bold=True)
            ws["A1"].alignment = Alignment(horizontal="center")
        else:
            for v_name, v_data in vessel_groups.items():
                # 1. Write the Vessel header block
                # Row 1
                ws.row_dimensions[current_row].height = 20
                
                c1_lbl = ws.cell(row=current_row, column=1, value="Name of vessel")
                c1_val = ws.cell(row=current_row, column=2, value=v_data["vessel_name"])
                c2_lbl = ws.cell(row=current_row, column=3, value="Vessel berthing date/time")
                
                b_dt = v_data["berthing_datetime"]
                b_dt_str = b_dt.strftime("%d.%m.%Y") if isinstance(b_dt, datetime) else (b_dt or "")
                c2_val = ws.cell(row=current_row, column=4, value=b_dt_str)

                # Row 2
                ws.row_dimensions[current_row + 1].height = 20
                c3_lbl = ws.cell(row=current_row + 1, column=1, value="Survey Qty")
                
                sq_val = float(v_data["survey_quantity"]) if v_data["survey_quantity"] is not None else ""
                c3_val = ws.cell(row=current_row + 1, column=2, value=sq_val)
                c4_lbl = ws.cell(row=current_row + 1, column=3, value="Vessel UNberthing date/time")
                
                s_dt = v_data["sailing_datetime"]
                s_dt_str = s_dt.strftime("%d.%m.%Y") if isinstance(s_dt, datetime) else (s_dt or "")
                c4_val = ws.cell(row=current_row + 1, column=4, value=s_dt_str)

                # Format Vessel Header Block cells
                for r in range(current_row, current_row + 2):
                    for col in range(1, 5):
                        cell = ws.cell(row=r, column=col)
                        cell.font = vessel_header_font
                        cell.fill = vessel_header_fill
                        cell.border = thin_border
                        if col in [1, 3]:
                            cell.alignment = left_align
                        else:
                            cell.alignment = center_align
                            if r == current_row + 1 and col == 2 and sq_val != "":
                                cell.number_format = "0.00"

                # 2. Leave 1 blank row
                current_row += 3
                
                # 3. Write Table Columns Header Row
                ws.row_dimensions[current_row].height = 32
                for col_idx, h_text in enumerate(headers_cols, 1):
                    cell = ws.cell(row=current_row, column=col_idx, value=h_text)
                    cell.font = table_header_font
                    cell.fill = table_header_fill
                    cell.alignment = center_align
                    cell.border = thin_border

                # 4. Write data rows
                start_data_row = current_row + 1
                for sl_no, r_data in enumerate(v_data["rows"], 1):
                    current_row += 1
                    ws.row_dimensions[current_row].height = 20
                    
                    row_values = [
                        sl_no,
                        r_data["party_name"],
                        r_data["gate_in_datetime"],
                        r_data["challan_invoice_no"],
                        r_data["vehicle_no"],
                        r_data["transporter_name"],
                        r_data["driver_name"],
                        r_data["driver_mob_no"],
                        r_data["outside_payment_slip"],
                        r_data["outside_gross_weight"],
                        r_data["outside_tare_weight"],
                        r_data["outside_net_weight"],
                        r_data["wbin_datetime"],
                        r_data["gross_weight"],
                        r_data["wbout_datetime"],
                        r_data["tare_weight"],
                        r_data["gate_out_datetime"],
                        r_data["net_weight"],
                        r_data["waiting_hours"],
                        r_data["cargo_start_datetime"],
                        r_data["compressor_no"],
                        r_data["cargo_end_datetime"]
                    ]

                    for col_idx, val in enumerate(row_values, 1):
                        cell = ws.cell(row=current_row, column=col_idx, value=val)
                        cell.font = Font(name="Calibri", size=11)
                        cell.border = thin_border
                        
                        # Alignment & format
                        if col_idx in [1, 3, 5, 8, 9, 13, 15, 17, 21]:
                            cell.alignment = center_align
                        elif col_idx in numeric_cols:
                            cell.alignment = right_align
                            if val is not None:
                                cell.number_format = numeric_cols[col_idx]
                        else:
                            cell.alignment = left_align

                        # Highlight Net Material Qty (Col 18)
                        if col_idx == 18:
                            cell.fill = yellow_fill
                            cell.font = net_font

                end_data_row = current_row

                # 5. Write Total Row
                current_row += 1
                ws.row_dimensions[current_row].height = 24
                
                # Merge A to Q for Total Label
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=17)
                total_label = ws.cell(row=current_row, column=1, value="Total")
                total_label.font = Font(name="Calibri", size=11, bold=True)
                total_label.alignment = right_align
                
                # Sum formula for Net Material Qty (Col 18)
                if end_data_row >= start_data_row:
                    sum_cell = ws.cell(row=current_row, column=18, value=f"=SUM(R{start_data_row}:R{end_data_row})")
                else:
                    sum_cell = ws.cell(row=current_row, column=18, value=0.0)
                sum_cell.font = Font(name="Calibri", size=11, bold=True, color="B25E00")
                sum_cell.alignment = right_align
                sum_cell.number_format = "0.00"
                sum_cell.fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

                # Apply borders to total row cells
                for col_idx in range(1, 23):
                    cell = ws.cell(row=current_row, column=col_idx)
                    cell.border = thin_border

                # 6. Leave 3 blank rows before the next vessel
                current_row += 4

        # ================= AUTO WIDTH =================
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            
            # Loop through all rows in this column to calculate width
            for cell in col:
                # Skip merged or title block cells that might skew column widths
                if cell.row < 4:
                    continue
                try:
                    if cell.value is not None:
                        val_str = str(cell.value)
                        # Handle newlines in header labels
                        lines = val_str.split('\n')
                        for line in lines:
                            max_length = max(max_length, len(line))
                except:
                    pass
            
            ws.column_dimensions[col_letter].width = max(max_length + 4, 12)

        # Set specific minimum widths for certain columns
        ws.column_dimensions["A"].width = 8   # SL. No.
        ws.column_dimensions["B"].width = 25  # Consignor
        ws.column_dimensions["C"].width = 24  # Gate entry dt
        ws.column_dimensions["D"].width = 18  # Challan
        ws.column_dimensions["E"].width = 15  # Vehicle
        ws.column_dimensions["F"].width = 22  # Transporter
        ws.column_dimensions["M"].width = 24  # IN Weighment Date
        ws.column_dimensions["O"].width = 24  # OUT Weighment Date
        ws.column_dimensions["Q"].width = 24  # GATE Out
        ws.column_dimensions["T"].width = 25  # Unloading start
        ws.column_dimensions["V"].width = 25  # Unloading complete

        wb.save(file_path)

        cursor.close()
        conn.close()

        return jsonify({
            "success": True,
            "download_url": f"/api/v1/export/download/{file_name}"
        }), 200

    except Exception as e:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500


def export_bills_report():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        query_param = request.args.get("query")
        start = request.args.get("start_date")
        end = request.args.get("end_date")
        bill_id = request.args.get("bill_id")

        conditions = []
        params = []

        if bill_id:
            conditions.append("bm.id = %s")
            params.append(bill_id)
        else:
            if query_param:
                conditions.append("(bm.voucher_number LIKE %s OR pm.party_name LIKE %s OR bm.narration LIKE %s)")
                like_val = f"%{query_param}%"
                params.extend([like_val, like_val, like_val])

            if start:
                conditions.append("bm.bill_date >= %s")
                params.append(start)

            if end:
                conditions.append("bm.bill_date <= %s")
                params.append(end)

        where = " AND ".join(conditions)
        if where:
            where = "WHERE " + where

        query = f"""
            SELECT 
                bm.voucher_number,
                DATE_FORMAT(bm.bill_date, '%%Y-%%m-%%d') as bill_date,
                pm.party_name,
                DATE_FORMAT(bm.period_start, '%%Y-%%m-%%d') as period_start,
                DATE_FORMAT(bm.period_end, '%%Y-%%m-%%d') as period_end,
                bm.bill_base_value,
                bm.cgst,
                bm.sgst,
                bm.round_off,
                bm.total_bill_value,
                bm.narration
            FROM bill_main bm
            LEFT JOIN party_masters pm ON pm.id = bm.party_id
            {where}
            ORDER BY bm.bill_date DESC, bm.id DESC
        """
        
        cursor.execute(query, tuple(params))
        cols = [c[0] for c in cursor.description]
        rows = cursor.fetchall()

        # Generate excel file
        if bill_id and len(rows) > 0:
            voucher_no = rows[0][0]
            safe_vch = "".join([c if c.isalnum() or c in ('-', '_') else '_' for c in voucher_no])
            file_name = f"BILL_{safe_vch}.xlsx"
        else:
            file_name = f"BILLS_REPORT_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
        file_path = os.path.join(EXPORT_FOLDER, file_name)

        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "BILLS_REPORT"
        ws.views.sheetView[0].showGridLines = True

        # Styles
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )
        
        title_font = Font(name="Calibri", size=14, bold=True, color="1B365D")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
        
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        total_font = Font(name="Calibri", size=11, bold=True, color="B25E00")

        headers = [
            "SL. No.",
            "Voucher No",
            "Bill Date",
            "Party Name",
            "Period Start",
            "Period End",
            "Base Value",
            "CGST",
            "SGST",
            "Round Off",
            "Total Value",
            "Narration"
        ]

        # Write Title
        ws.merge_cells("A1:L1")
        ws["A1"] = "ALL GENERATED BILLS REPORT"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        ws.row_dimensions[1].height = 25

        # Write filter summary in row 2
        ws.merge_cells("A2:L2")
        filter_summary = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        if start or end:
            filter_summary += f" | Date Range: {start or 'Any'} to {end or 'Any'}"
        if query_param:
            filter_summary += f" | Search filter: '{query_param}'"
        ws["A2"] = filter_summary
        ws["A2"].font = Font(name="Calibri", size=10, italic=True)
        ws["A2"].alignment = center_align
        ws.row_dimensions[2].height = 20

        # Leave row 3 blank
        ws.row_dimensions[3].height = 10

        # Write Headers (Row 4)
        ws.row_dimensions[4].height = 28
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # Write Data
        start_row = 5
        for idx, row in enumerate(rows, 1):
            r_idx = start_row + idx - 1
            ws.row_dimensions[r_idx].height = 20
            
            # Map values
            vch_no, bill_dt, party, p_start, p_end, base, cgst, sgst, round_off, total, narr = row
            
            row_values = [
                idx,
                vch_no,
                bill_dt,
                party or "—",
                p_start or "—",
                p_end or "—",
                float(base) if base is not None else 0.00,
                float(cgst) if cgst is not None else 0.00,
                float(sgst) if sgst is not None else 0.00,
                float(round_off) if round_off is not None else 0.00,
                float(total) if total is not None else 0.00,
                narr or ""
            ]
            
            for col_idx, val in enumerate(row_values, 1):
                cell = ws.cell(row=r_idx, column=col_idx, value=val)
                cell.font = Font(name="Calibri", size=11)
                cell.border = thin_border
                
                # Alignments and number formats
                if col_idx in [1, 2, 3, 5, 6]:
                    cell.alignment = center_align
                elif col_idx in [7, 8, 9, 10, 11]:
                    cell.alignment = right_align
                    cell.number_format = "0.00"
                else:
                    cell.alignment = left_align

        end_row = start_row + len(rows) - 1

        # Write Totals Row
        total_row = end_row + 1
        if len(rows) > 0:
            ws.row_dimensions[total_row].height = 24
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=6)
            t_lbl = ws.cell(row=total_row, column=1, value="Total")
            t_lbl.font = Font(name="Calibri", size=11, bold=True)
            t_lbl.alignment = right_align
            
            # Sum columns
            sum_cols = [7, 8, 9, 10, 11] # Base, CGST, SGST, Round Off, Total
            col_letters = ["G", "H", "I", "J", "K"]
            
            for col_idx, col_letter in zip(sum_cols, col_letters):
                cell = ws.cell(row=total_row, column=col_idx, value=f"=SUM({col_letter}{start_row}:{col_letter}{end_row})")
                cell.font = total_font
                cell.fill = total_fill
                cell.alignment = right_align
                cell.number_format = "0.00"
                
            # Apply border to the merged cells and remaining cells in total row
            for col_idx in range(1, 13):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.border = thin_border
                if col_idx in [1, 2, 3, 4, 5, 6]:
                    cell.fill = total_fill

        # Auto width columns
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.row < 4:
                    continue
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max(max_length + 4, 12)

        # Set specific minimums
        ws.column_dimensions["A"].width = 8   # SL. No.
        ws.column_dimensions["B"].width = 18  # Voucher No
        ws.column_dimensions["C"].width = 14  # Bill Date
        ws.column_dimensions["D"].width = 25  # Party Name
        ws.column_dimensions["L"].width = 30  # Narration

        wb.save(file_path)
        cursor.close()
        conn.close()

        return jsonify({
            "success": True,
            "download_url": f"/api/v1/export/download/{file_name}"
        }), 200

    except Exception as e:
        if 'cursor' in locals() and cursor:
            cursor.close()
        if 'conn' in locals() and conn:
            conn.close()
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500